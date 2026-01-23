# excel_reports.py
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
import database
from contextlib import contextmanager
from datetime import datetime
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

router = APIRouter()


@contextmanager
def get_db():
    conn = database.get_db_connection()
    try:
        yield conn
    finally:
        if conn:
            conn.close()


# Excel отчет по сеансам за определенную дату
@router.get("/sessions-excel-report")
async def sessions_excel_report(date: str = None):
    try:
        if not date:
            # Если дата не указана, используем сегодняшнюю
            date = datetime.now().strftime("%Y-%m-%d")

        # Форматируем дату для отображения
        formatted_date = datetime.strptime(date, '%Y-%m-%d').strftime('%d.%m.%Y')

        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT s.*, m.movie_title 
                    FROM session s 
                    JOIN movie m ON s.id_movie = m.id_movie 
                    WHERE s.data_session = %s 
                    ORDER BY s.start
                """, (date,))
                sessions_data = cur.fetchall()

                sessions = []
                for session in sessions_data:
                    sessions.append({
                        "ID сеанса": session["id_session"],
                        "Фильм": session["movie_title"],
                        "Дата": session["data_session"].strftime("%d.%m.%Y"),
                        "Начало": str(session["start"]),
                        "Окончание": str(session["final"]),
                        "Зал": session["hall"]
                    })

        # Создаем Excel файл в памяти
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Сеансы"

        # Стиль для границ
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        if not sessions:
            # Если нет сеансов на выбранную дату
            worksheet.merge_cells('A1:F3')
            cell = worksheet['A1']
            cell.value = f"Расписание сеансов на {formatted_date}\n\nНа выбранную дату нет сеансов"
            cell.font = Font(size=14, bold=True, color="FF0000")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

            # Устанавливаем высоту строки
            worksheet.row_dimensions[1].height = 80

        else:
            # Добавляем заголовок отчета
            worksheet.merge_cells('A1:F1')
            title_cell = worksheet['A1']
            title_cell.value = f"Расписание сеансов на {formatted_date}"
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')

            # Пустая строка после заголовка
            worksheet.row_dimensions[2].height = 15

            # Создаем DataFrame
            df = pd.DataFrame(sessions)

            # Добавляем данные из DataFrame (начинаем с 3-й строки)
            start_row = 3
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border

                    # Форматируем заголовки
                    if r_idx == start_row:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        # Для данных центрируем все колонки кроме "Фильм"
                        if c_idx != 2:  # Не центрируем колонку "Фильм"
                            cell.alignment = Alignment(horizontal='center')

            # Настраиваем ширину колонок
            worksheet.column_dimensions['A'].width = 12  # ID сеанса
            worksheet.column_dimensions['B'].width = 35  # Фильм
            worksheet.column_dimensions['C'].width = 12  # Дата
            worksheet.column_dimensions['D'].width = 10  # Начало
            worksheet.column_dimensions['E'].width = 10  # Окончание
            worksheet.column_dimensions['F'].width = 8  # Зал

            # ДобаВЛЯЕМ СТРОКУ "Всего сеансов" как часть таблицы
            total_row = start_row + len(sessions) + 1
            worksheet.merge_cells(f'A{total_row}:F{total_row}')
            total_cell = worksheet.cell(row=total_row, column=1, value=f"Всего сеансов: {len(sessions)}")
            total_cell.font = Font(bold=True)
            total_cell.alignment = Alignment(horizontal='center')
            total_cell.border = thin_border

            # Добавляем пустые строки после таблицы
            empty_row = total_row + 2

            # Добавляем подпись и дату (не жирным)
            worksheet.cell(row=empty_row, column=1, value="Подпись: ___________________")
            worksheet.cell(row=empty_row, column=1).font = Font(size=12)  # Убрали bold=True

            worksheet.cell(row=empty_row + 1, column=1, value="Дата: ___________________")
            worksheet.cell(row=empty_row + 1, column=1).font = Font(size=12)  # Убрали bold=True

        # Сохраняем workbook в поток
        workbook.save(output)
        output.seek(0)

        filename = f"raspisanie_seansov_{date}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        return {"error": f"Ошибка генерации Excel отчета: {str(e)}"}


# Получить список доступных дат с сеансами
@router.get("/api/available-dates")
async def get_available_dates():
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT DISTINCT data_session 
                    FROM session 
                    WHERE data_session >= CURRENT_DATE 
                    ORDER BY data_session
                """)
                dates_data = cur.fetchall()

                dates = []
                for date_row in dates_data:
                    dates.append(date_row["data_session"].strftime("%Y-%m-%d"))

        return dates
    except Exception as e:
        return {"error": str(e)}