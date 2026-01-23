# modules/reports.py
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta
import database
from contextlib import contextmanager
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from modules.auth import get_current_user, check_permission

router = APIRouter()


@contextmanager
def get_db():
    conn = database.get_db_connection()
    try:
        yield conn
    finally:
        if conn:
            conn.close()


# ========== ОТЧЕТЫ ==========
@router.get("/api/reports/daily-sales")
async def daily_sales_report(date: str = None, current_user: dict = Depends(get_current_user)):
    """Отчет о продажах за день"""
    if not check_permission(current_user, "reports:generate"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    try:
        if not date:
            date = datetime.now().strftime("%Y-%m-%d")

        with get_db() as conn:
            with conn.cursor() as cur:
                # Получаем данные о продажах за день
                cur.execute("""
                    SELECT 
                        t.id_ticket,
                        m.movie_title,
                        s.data_session,
                        s.start,
                        s.hall,
                        t.category,
                        t.price,
                        t.place,
                        t.row,
                        t.sold,
                        t.created_at
                    FROM tickets t
                    JOIN movie m ON t.id_movie = m.id_movie
                    JOIN session s ON t.id_session = s.id_session
                    WHERE DATE(t.created_at) = %s
                    ORDER BY t.created_at
                """, (date,))

                sales_data = cur.fetchall()

                # Получаем сводную статистику
                cur.execute("""
                    SELECT 
                        COUNT(*) as total_tickets,
                        SUM(CASE WHEN sold = 'sold' THEN 1 ELSE 0 END) as sold_tickets,
                        SUM(CASE WHEN sold = 'free' THEN 1 ELSE 0 END) as free_tickets,
                        COALESCE(SUM(CASE WHEN sold = 'sold' THEN price ELSE 0 END), 0) as total_revenue
                    FROM tickets
                    WHERE DATE(created_at) = %s
                """, (date,))

                stats = cur.fetchone()

        # Формируем Excel отчет
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Продажи {date}"

        # Стиль для границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовок отчета
        worksheet.merge_cells('A1:K1')
        title_cell = worksheet['A1']
        title_cell.value = f"Отчет о продажах за {date}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Статистика
        worksheet['A3'] = "Всего билетов:"
        worksheet['B3'] = stats['total_tickets'] or 0
        worksheet['A4'] = "Проданных билетов:"
        worksheet['B4'] = stats['sold_tickets'] or 0
        worksheet['A5'] = "Свободных билетов:"
        worksheet['B5'] = stats['free_tickets'] or 0
        worksheet['A6'] = "Общая выручка:"
        worksheet['B6'] = float(stats['total_revenue'] or 0)

        # Заголовки таблицы
        headers = [
            "ID билета", "Фильм", "Дата сеанса", "Время", "Зал",
            "Категория", "Цена", "Место", "Ряд", "Статус", "Дата создания"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=8, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Данные
        row_idx = 9
        for sale in sales_data:
            worksheet.cell(row=row_idx, column=1, value=sale['id_ticket'])
            worksheet.cell(row=row_idx, column=2, value=sale['movie_title'])
            worksheet.cell(row=row_idx, column=3,
                           value=sale['data_session'].strftime("%d.%m.%Y") if sale['data_session'] else "")
            worksheet.cell(row=row_idx, column=4, value=str(sale['start'])[:5] if sale['start'] else "")
            worksheet.cell(row=row_idx, column=5, value=sale['hall'])
            worksheet.cell(row=row_idx, column=6, value=sale['category'])
            worksheet.cell(row=row_idx, column=7, value=float(sale['price']) if sale['price'] else 0)
            worksheet.cell(row=row_idx, column=8, value=sale['place'])
            worksheet.cell(row=row_idx, column=9, value=sale['row'])
            worksheet.cell(row=row_idx, column=10, value=sale['sold'])
            worksheet.cell(row=row_idx, column=11,
                           value=sale['created_at'].strftime("%d.%m.%Y %H:%M") if sale['created_at'] else "")

            # Добавляем границы
            for col in range(1, 12):
                worksheet.cell(row=row_idx, column=col).border = thin_border

            row_idx += 1

        # Автоматическая ширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем workbook
        workbook.save(output)
        output.seek(0)

        filename = f"daily_sales_{date}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации отчета: {str(e)}")


@router.get("/api/reports/films-popularity")
async def films_popularity_report(current_user: dict = Depends(get_current_user)):
    """Отчет о популярности фильмов"""
    if not check_permission(current_user, "reports:generate"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                # Получаем статистику по фильмам
                cur.execute("""
                    SELECT 
                        m.movie_title,
                        COUNT(t.id_ticket) as total_tickets,
                        SUM(CASE WHEN t.sold = 'sold' THEN 1 ELSE 0 END) as sold_tickets,
                        COUNT(DISTINCT s.id_session) as total_sessions,
                        COALESCE(SUM(CASE WHEN t.sold = 'sold' THEN t.price ELSE 0 END), 0) as total_revenue,
                        COALESCE(AVG(CASE WHEN t.sold = 'sold' THEN t.price ELSE NULL END), 0) as avg_price
                    FROM movie m
                    LEFT JOIN session s ON m.id_movie = s.id_movie
                    LEFT JOIN tickets t ON s.id_session = t.id_session
                    GROUP BY m.movie_title
                    ORDER BY sold_tickets DESC
                """)

                films_data = cur.fetchall()

        # Формируем отчет
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Популярность фильмов"

        # Стиль для границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовок отчета
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet['A1']
        title_cell.value = "Отчет о популярности фильмов"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Заголовки таблицы
        headers = ["Фильм", "Всего билетов", "Проданных", "Сеансов", "Выручка", "Ср. цена"]

        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Данные
        row_idx = 4
        for film in films_data:
            worksheet.cell(row=row_idx, column=1, value=film['movie_title'])
            worksheet.cell(row=row_idx, column=2, value=film['total_tickets'] or 0)
            worksheet.cell(row=row_idx, column=3, value=film['sold_tickets'] or 0)
            worksheet.cell(row=row_idx, column=4, value=film['total_sessions'] or 0)
            worksheet.cell(row=row_idx, column=5, value=float(film['total_revenue'] or 0))
            worksheet.cell(row=row_idx, column=6, value=float(film['avg_price'] or 0))

            # Добавляем границы
            for col in range(1, 7):
                worksheet.cell(row=row_idx, column=col).border = thin_border

            row_idx += 1

        # Автоматическая ширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем workbook
        workbook.save(output)
        output.seek(0)

        filename = "films_popularity.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации отчета: {str(e)}")


@router.get("/api/reports/monthly")
async def monthly_report(year: int = None, month: int = None, current_user: dict = Depends(get_current_user)):
    """Ежемесячный отчет"""
    if not check_permission(current_user, "reports:generate"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    try:
        if not year or not month:
            now = datetime.now()
            year = now.year
            month = now.month

        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"

        with get_db() as conn:
            with conn.cursor() as cur:
                # Получаем статистику за месяц
                cur.execute("""
                    SELECT 
                        DATE(t.created_at) as sale_date,
                        COUNT(*) as total_tickets,
                        SUM(CASE WHEN t.sold = 'sold' THEN 1 ELSE 0 END) as sold_tickets,
                        COALESCE(SUM(CASE WHEN t.sold = 'sold' THEN t.price ELSE 0 END), 0) as daily_revenue
                    FROM tickets t
                    WHERE t.created_at >= %s AND t.created_at < %s
                    GROUP BY DATE(t.created_at)
                    ORDER BY sale_date
                """, (start_date, end_date))

                daily_stats = cur.fetchall()

                # Общая статистика за месяц
                cur.execute("""
                    SELECT 
                        COUNT(*) as total_tickets,
                        SUM(CASE WHEN sold = 'sold' THEN 1 ELSE 0 END) as sold_tickets,
                        COALESCE(SUM(CASE WHEN sold = 'sold' THEN price ELSE 0 END), 0) as total_revenue,
                        COALESCE(AVG(CASE WHEN sold = 'sold' THEN price ELSE NULL END), 0) as avg_ticket_price
                    FROM tickets
                    WHERE created_at >= %s AND created_at < %s
                """, (start_date, end_date))

                monthly_stats = cur.fetchone()

        # Формируем отчет
        output = io.BytesIO()
        workbook = Workbook()

        # Лист с ежедневной статистикой
        daily_ws = workbook.active
        daily_ws.title = "Ежедневная статистика"

        # Заголовок
        daily_ws.merge_cells('A1:D1')
        title_cell = daily_ws['A1']
        title_cell.value = f"Ежемесячный отчет за {month:02d}.{year}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Общая статистика
        daily_ws['A3'] = "Общая статистика за месяц:"
        daily_ws['A4'] = "Всего билетов:"
        daily_ws['B4'] = monthly_stats['total_tickets'] or 0
        daily_ws['A5'] = "Проданных билетов:"
        daily_ws['B5'] = monthly_stats['sold_tickets'] or 0
        daily_ws['A6'] = "Общая выручка:"
        daily_ws['B6'] = float(monthly_stats['total_revenue'] or 0)
        daily_ws['A7'] = "Средняя цена билета:"
        daily_ws['B7'] = float(monthly_stats['avg_ticket_price'] or 0)

        # Заголовки ежедневной статистики
        headers = ["Дата", "Всего билетов", "Проданных", "Выручка за день"]
        start_row = 10

        for col_idx, header in enumerate(headers, 1):
            cell = daily_ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Данные ежедневной статистики
        row_idx = start_row + 1
        total_monthly_revenue = 0

        for day in daily_stats:
            daily_ws.cell(row=row_idx, column=1,
                          value=day['sale_date'].strftime("%d.%m.%Y") if day['sale_date'] else "")
            daily_ws.cell(row=row_idx, column=2, value=day['total_tickets'] or 0)
            daily_ws.cell(row=row_idx, column=3, value=day['sold_tickets'] or 0)
            daily_ws.cell(row=row_idx, column=4, value=float(day['daily_revenue'] or 0))

            total_monthly_revenue += float(day['daily_revenue'] or 0)
            row_idx += 1

        # Итоговая строка
        daily_ws.cell(row=row_idx, column=1, value="ИТОГО:")
        daily_ws.cell(row=row_idx, column=2, value=monthly_stats['total_tickets'] or 0)
        daily_ws.cell(row=row_idx, column=3, value=monthly_stats['sold_tickets'] or 0)
        daily_ws.cell(row=row_idx, column=4, value=float(monthly_stats['total_revenue'] or 0))

        # Жирный шрифт для итогов
        for col in range(1, 5):
            daily_ws.cell(row=row_idx, column=col).font = Font(bold=True)

        # Автоматическая ширина колонок
        for column in daily_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            daily_ws.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем workbook
        workbook.save(output)
        output.seek(0)

        filename = f"monthly_report_{year}_{month:02d}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации отчета: {str(e)}")


# Статистика для дашборда
@router.get("/api/reports/dashboard-stats")
async def dashboard_stats(current_user: dict = Depends(get_current_user)):
    """Статистика для дашборда"""
    if not check_permission(current_user, "reports:basic"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                # Общая статистика
                cur.execute("SELECT COUNT(*) as total FROM movie")
                total_films = cur.fetchone()['total']

                cur.execute("SELECT COUNT(*) as total FROM session WHERE data_session = CURRENT_DATE")
                today_sessions = cur.fetchone()['total']

                cur.execute("""
                    SELECT COUNT(*) as total 
                    FROM tickets 
                    WHERE sold = 'sold' AND DATE(created_at) = CURRENT_DATE
                """)
                today_sales = cur.fetchone()['total']

                cur.execute("""
                    SELECT COALESCE(SUM(price), 0) as total 
                    FROM tickets 
                    WHERE sold = 'sold' AND DATE(created_at) = CURRENT_DATE
                """)
                today_revenue = cur.fetchone()['total'] or 0

                # Популярные фильмы
                cur.execute("""
                    SELECT m.movie_title, COUNT(t.id_ticket) as ticket_count
                    FROM movie m
                    JOIN session s ON m.id_movie = s.id_movie
                    JOIN tickets t ON s.id_session = t.id_session
                    WHERE t.sold = 'sold'
                    GROUP BY m.movie_title
                    ORDER BY ticket_count DESC
                    LIMIT 5
                """)
                popular_films = cur.fetchall()

        return {
            "stats": {
                "total_films": total_films,
                "today_sessions": today_sessions,
                "today_sales": today_sales,
                "today_revenue": float(today_revenue)
            },
            "popular_films": [
                {"film": film["movie_title"], "tickets": film["ticket_count"]}
                for film in popular_films
            ]
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения статистики: {str(e)}")